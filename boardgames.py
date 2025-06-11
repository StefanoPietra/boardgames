#! python3
# boardgames.py v2.0 - Update board game prices from Zatu and BoardGamePrices,
# saving and formatting them in a new sheet of the Excel file
# v2.0 - check prices from Zatu website too

import logging, openpyxl, datetime, requests, bs4, re
from openpyxl.styles import PatternFill, Font

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", datefmt="%H:%M:%S")
logging.disable(logging.INFO)

### PARAMETERS
excelFile = "Data\Boardgames.xlsx"
zatuWebsite = "https://www.board-game.co.uk/product/"
BGPWebsite = "https://boardgameprices.co.uk/item/show/"
priceRegex = re.compile(r"£(\d+\.\d+)")    # regular expression for prices (£##.##)
zatuDeliveryFee = 2.99      # added to all Zatu prices
priceChange = 20  # percentage change over which the price is considered different

def findColumn(header):
    """Return number of the column with the given HEADER"""
    for cellObj in list(latestSheet.rows)[0]:   # Loop over row1
        if cellObj.value == header:
            columnNumber = cellObj.column
            logging.debug("Found " + header + " column: " + str(columnNumber))
            return(columnNumber)

def fillCell(newCell, oldCell):
    """Compare prices between NEWCELL and OLDCELL and set filling color of 
    NEWCELL to red, green or transparent"""
    redFill = PatternFill(fill_type="solid", start_color="F2DCDB", end_color="F2DCDB")
    greenFill = PatternFill(fill_type="solid", start_color="EBF1DE", end_color="EBF1DE")
    noFill = PatternFill(fill_type="none")
    if str(newCell.value).startswith("=SUM"):    # in case the cell contains a SUM formula
        firstNum = float(re.findall("=SUM\((\d+.\d+)", newCell.value)[0])
        secondNum = float(re.findall("\+(\d+.\d+)", newCell.value)[0])
        newPrice = firstNum + secondNum
    else:
        newPrice = newCell.value
    if str(oldCell.value).startswith("=SUM"):    # in case the cell contains a SUM formula
        firstNum = float(re.findall("=SUM\((\d+.\d+)", oldCell.value)[0])
        secondNum = float(re.findall("\+(\d+.\d+)", oldCell.value)[0])
        oldPrice = firstNum + secondNum
    else:
        oldPrice = oldCell.value
    priceDifference = (newPrice - oldPrice)*100 / oldPrice   # price change in % of the old price
    if priceDifference > priceChange:
        logging.info("New price is HIGHER: filling cell red")
        newCell.fill = redFill
    elif priceDifference < -priceChange:
        logging.info("New price is LOWER: filling cell green")
        newCell.fill = greenFill
    else:
        newCell.fill = noFill

logging.info("Starting program")

### Open Excel file and create a new sheet
logging.info("Opening Excel file: " + excelFile)
workbook = openpyxl.load_workbook(excelFile)
latestSheet = workbook[workbook.sheetnames[-1]]
zatuAddressColumn = findColumn("Zatu address")
BGPAddressColumn = findColumn("BoardGamePrices address")
zatuFullColumn = findColumn("Zatu Full")
zatuScontatoColumn = findColumn("Zatu Scontato")
otherBestColumn = findColumn("Other best")
# Numbers associated to columns that will be read and written to
logging.info("Columns of interest: " \
              + ", ".join([str(x) for x in [zatuAddressColumn, BGPAddressColumn, zatuFullColumn, zatuScontatoColumn, otherBestColumn]]))

logging.info("Duplicating the latest sheet")
newSheet = workbook.copy_worksheet(latestSheet)

logging.debug("Deciding name for the new sheet")
dt = datetime.datetime.now()
newSheetName = dt.strftime("%Y-%m")     # year and month, 2024-07
if newSheetName == latestSheet.title :
    newSheetName = newSheetName+"(1)"       # if the month already exists -> 2024-07(1)

logging.info("Renaming the new sheet as: " + newSheetName)
newSheet.title = newSheetName

### Loop over the games
for row in range(1, newSheet.max_row):
    if newSheet.cell(row=row, column=1).value == None:  # skip empty rows
        continue
    game = newSheet.cell(row=row, column=1).value

    ### Visit Zatu and scrape the game's prices and availability
    logging.debug("Building URL for " + game + " on Zatu")
    zatuUrl = zatuWebsite + newSheet.cell(row=row, column=zatuAddressColumn).value

    logging.info("Getting web page:\n           " + zatuUrl)
    try:    # try to download page
        zatuRes = requests.get(zatuUrl)
        zatuRes.raise_for_status()
    except:
        logging.error("Impossible to retrieve page for " + game + " on Zatu")

    if zatuRes.status_code == requests.codes.ok:    # if download worked
        zatuGamePage = bs4.BeautifulSoup(zatuRes.text, "html.parser")
        
        try:    # try to extract prices
            logging.info("Parsing page to find prices")
            zatuFullPriceElem = zatuGamePage.select(".zg-single-price-box-was")
            zatuScontatoPriceElem = zatuGamePage.select(".zg-single-price-box-now")
            if len(zatuFullPriceElem) < 1:  # in case there is no discount and there is only a "NOW" price
                zatuFullPriceElem = zatuScontatoPriceElem
            zatuFullPrice = float(priceRegex.findall(zatuFullPriceElem[0].text)[0])   # transform price into a number
            logging.info("Full price for " + game + " on Zatu is " + str(round(zatuFullPrice+zatuDeliveryFee,2)))
            newSheet.cell(row=row, column=zatuFullColumn).value = "=SUM("+str(zatuFullPrice)+"+"+str(zatuDeliveryFee)+")"
            logging.debug("Comparing price with the previous sheet")
            fillCell(newSheet.cell(row=row, column=zatuFullColumn), latestSheet.cell(row=row, column=zatuFullColumn))

            zatuScontatoPrice = float(priceRegex.findall(zatuScontatoPriceElem[0].text)[0])     # transform price into a number
            logging.info("Scontato price for " + game + " on Zatu is " + str(round(zatuScontatoPrice+zatuDeliveryFee,2)))
            newSheet.cell(row=row, column=zatuScontatoColumn).value = "=SUM("+str(zatuScontatoPrice)+"+"+str(zatuDeliveryFee)+")"
            logging.debug("Comparing price with the previous sheet")
            fillCell(newSheet.cell(row=row, column=zatuScontatoColumn), latestSheet.cell(row=row, column=zatuScontatoColumn))
        except:
            logging.error("Impossible to extract prices for " + game + " on Zatu")
     
        try:    # try to check availability
            logging.info("Checking availability")   # based on text on the orange button
            zatuAvailabilityElem = zatuGamePage.select("button")
            zatuAvailability = zatuAvailabilityElem[4].text
            logging.debug(zatuAvailability)
            if zatuAvailability == "Add to basket":
                newSheet.cell(row=row, column=zatuFullColumn).font = Font(color="000000")
                newSheet.cell(row=row, column=zatuScontatoColumn).font = Font(color="000000")
            elif zatuAvailability == "Notify Me" or zatuAvailability == "Place Backorder":
                logging.info(game + " is NOT available on Zatu")
                newSheet.cell(row=row, column=zatuFullColumn).font = Font(color="FF0000")
                newSheet.cell(row=row, column=zatuScontatoColumn).font = Font(color="FF0000")
            else:
                logging.error("Unexpected availability info for " + game + " on Zatu")
        except:
            logging.error("Impossible to retrieve availability info for " + game + " on Zatu")

    ### Visit BoardGamePrices and scrape the game's price and availability
    logging.debug("Building URL for " + game + " on BoardGamePrices")
    BGPUrl = BGPWebsite + newSheet.cell(row=row, column=BGPAddressColumn).value

    logging.info("Getting web page\n           " + BGPUrl)
    try:    # try to download page
        BGPRes = requests.get(BGPUrl)
        BGPRes.raise_for_status()
    except:
        logging.error("Impossible to retrieve page for " + game + " on BoardGamePrices")

    if BGPRes.status_code == requests.codes.ok:    # if download worked
        BGPGamePage = bs4.BeautifulSoup(BGPRes.text, "html.parser")
        
        try:    # try to extract price
            logging.info("Parsing page to find cheapest price") # (with shipping)
            otherBestPriceElem = BGPGamePage.select("#vendorlist > div:nth-child(8) > div > div.total.grand-total")
            otherBestPrice = float(priceRegex.findall(otherBestPriceElem[0].text)[0])    # transform price into a number
            logging.info("Best price for " + game + " on BoardGamePrices is " + str(otherBestPrice))
            newSheet.cell(row=row, column=otherBestColumn).value = otherBestPrice
            logging.debug("Comparing price with the previous sheet")
            fillCell(newSheet.cell(row=row, column=otherBestColumn), latestSheet.cell(row=row, column=otherBestColumn))
        except:
            logging.error("Impossible to extract price for " + game + " on BoardGamePrices")

        try:    # try to check availability
            logging.info("Checking availability")   # the green light
            BGPAvailabilityElem = BGPGamePage.select("#vendorlist > div:nth-child(7) > div.infocontainer.multicell > div.vendorstock > span")
            BGPAvailability = BGPAvailabilityElem[0].text
            logging.debug(BGPAvailability)
            if BGPAvailability == "Yes":
                newSheet.cell(row=row, column=otherBestColumn).font = Font(color="000000")
            elif BGPAvailability == "No":
                logging.info(game + " is NOT available on BoardGamePrices")
                newSheet.cell(row=row, column=otherBestColumn).font = Font(color="FF0000")
            else:
                logging.error("Unexpected availability info for " + game + " on BoardGamePrices")
        except:
            logging.error("Impossible to retrieve availability info for " + game + " on BoardGamePrices")

logging.info("Saving the Excel file")
workbook.active.views.sheetView[0].tabSelected = False  # disable the current active sheet
workbook.active = newSheet  # set new active sheet
workbook.save(excelFile)
